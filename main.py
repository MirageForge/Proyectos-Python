import tkinter as tk
import re
import os
from tkinter import messagebox, Label
from openpyxl import Workbook, load_workbook

nombre_archivo = 'Datos.xlsx'

if os.path.exists(nombre_archivo):
    wb = load_workbook(nombre_archivo)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    wb.save(nombre_archivo)

wb = Workbook()
ws = wb.active
ws.append(["Nombre", "Edad", "Direcccion", "Email", "Telefono",])

def guardar_datos():
    nombre=entry_nombre.get()
    edad=entry_edad.get()
    direccion=entry_direccion.get()
    email=entry_email.get()
    telefono=entry_telefono.get()

    if not nombre or not edad or not direccion or not email or not telefono:
        messagebox.showwarning(title="Advertencia", message="Todos los campos son obligatorios")
        return
    try:
        edad= int(edad)
        telefono = int(telefono)
    except ValueError:
        messagebox.showwarning(title="Error", message="Edad y Telefono deben de ser numeros")
        return

    if not re.match(r"[^@]+@[^@]+\.[^@]+", email):
        messagebox.showwarning(title="Advertencia", message="El correo electrónico no es válido")

    ws.append([nombre, edad, direccion, email, telefono])
    wb.save(nombre_archivo)
    messagebox.showinfo(title="Informacion", message="Datos guardados correctamente")

    entry_nombre.delete(0, tk.END)
    entry_edad.delete(0, tk.END)
    entry_direccion.delete(0, tk.END)
    entry_email.delete(0, tk.END)
    entry_telefono.delete(0, tk.END)

root = tk.Tk()
root.title("Formulario de Entrada de Datos")
root.configure(bg="#4B6587")
label_style =  {"bg": '#4B6587', "fg": "white"}
entry_style =  {"bg": '#D3D3D3', "fg": "black"}

label_nombre = tk.Label(root, text="Nombre", **label_style)
label_nombre.grid(row=0, column=0, padx=10, pady=5)
entry_nombre = tk.Entry(root, **entry_style)
entry_nombre.grid(row=0, column=1, padx=10, pady=5)

label_edad = tk.Label(root, text="Edad", **label_style)
label_edad.grid(row=1, column=0, padx=10, pady=5)
entry_edad = tk.Entry(root, **entry_style)
entry_edad.grid(row=1, column=1, padx=10, pady=5)

label_direccion = tk.Label(root, text="Direcccion", **label_style)
label_direccion.grid(row=2, column=0, padx=10, pady=5)
entry_direccion = tk.Entry(root, **entry_style)
entry_direccion.grid(row=2, column=1, padx=10, pady=5)

label_email = tk.Label(root, text="Email", **label_style)
label_email.grid(row=3, column=0, padx=10, pady=5)
entry_email = tk.Entry(root, **entry_style)
entry_email.grid(row=3, column=1, padx=10, pady=5)

label_telefono = tk.Label(root, text="Telefono", **label_style)
label_telefono.grid(row=4, column=0, padx=10, pady=5)
entry_telefono = tk.Entry(root, **entry_style)
entry_telefono.grid(row=4, column=1, padx=10, pady=5)

boton_guardar = tk.Button(root, text="Guardar", command=guardar_datos, bg='#6D8299', fg='white')
boton_guardar.grid(row=5, column=0, columnspan=2, padx=10, pady=5)
root.mainloop()
